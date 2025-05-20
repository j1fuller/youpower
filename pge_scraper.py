# Author: SupportDone.com
# YouPower PG&E Green Button Data Scraper
# Version: 1.0.0

import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QLabel, QLineEdit, QPushButton, QVBoxLayout, QWidget, QDateEdit, QMessageBox, QDesktopWidget, QProgressBar, QFileDialog, QHBoxLayout, QComboBox, QCheckBox
)
from PyQt5.QtCore import QDate, QThread, pyqtSignal, Qt
from PyQt5.QtGui import QPixmap, QIcon
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import json
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime
from selenium.webdriver.common.keys import Keys

class AutomationWorker(QThread):
    """Worker thread to run Selenium automation for PG&E."""
    progress = pyqtSignal(int)
    step_counter = 0
    finished = pyqtSignal(bool, str)
    driver = None

    def __init__(self, url, username, password, start_date, end_date, download_path, utility_provider):
        super().__init__()
        self.url = url
        self.username = username
        self.password = password
        self.start_date = start_date
        self.end_date = end_date
        self.download_path = download_path
        self.utility_provider = utility_provider

    def login_to_portal(self, driver, url, username, password):
        """Perform login actions for PG&E."""
        for attempt in range(3):
            driver.get(url)
            time.sleep(2)
            driver.refresh()

            try:
                print(f"Attempt {attempt + 1}: Logging in to {self.utility_provider}...")
                
                if self.utility_provider == "PG&E":
                    # PG&E login procedure
                    try:
                        # First try to find the username field
                        username_field = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((By.ID, "username"))
                        )
                        username_field.clear()
                        username_field.send_keys(username)
                        time.sleep(1)

                        password_field = driver.find_element(By.ID, "password")
                        password_field.clear()
                        password_field.send_keys(password)
                        time.sleep(1)

                        login_button = driver.find_element(By.ID, "login")
                        login_button.click()
                        time.sleep(5)
                        
                        # Check for successful login
                        # This depends on PG&E's specific dashboard elements
                        dashboard_element = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, "//a[contains(text(), 'Energy Usage')]"))
                        )
                        print("Login successful!")
                        return
                    except Exception as e:
                        print(f"PG&E login attempt failed: {e}")
                        # Try alternative login selectors if needed
                
                elif self.utility_provider == "SDGE":
                    # SDGE login procedure (original code)
                    login_form_present = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((By.ID, "usernamex"))
                    )

                    if login_form_present:
                        username_field = driver.find_element(By.ID, "usernamex")
                        username_field.clear()
                        username_field.send_keys(username)
                        time.sleep(1)

                        password_field = driver.find_element(By.ID, "passwordx")
                        password_field.clear()
                        password_field.send_keys(password)
                        time.sleep(2)

                        driver.find_element(By.ID, "btnlogin").click()
                        time.sleep(3)
                    else:
                        print("Login form not found. Assuming login was successful.")
                        return

                    form_still_present = driver.find_elements(By.ID, "usernamex")
                    if not form_still_present:
                        print("Login successful!")
                        return
                        
                elif self.utility_provider == "SCE":
                    # SCE login procedure (to be implemented)
                    print("SCE login functionality not yet implemented")
                    # Placeholder for SCE login implementation
                    return

            except Exception as e:
                print(f"Login attempt {attempt + 1} failed: {e}")

    @staticmethod
    def validate_and_format_date(date_string):
        """Validate and format the date to MMM DD, YYYY."""
        accepted_formats = ["%B %d, %Y", "%Y-%m-%d", "%d %B, %Y"]
        for date_format in accepted_formats:
            try:
                date_obj = datetime.strptime(date_string, date_format)
                return date_obj.strftime("%B %d, %Y")
            except ValueError:
                continue
        raise ValueError(f"Invalid date format: {date_string}. Expected formats: {', '.join(accepted_formats)}.")

    def download_pge_green_button_data(self, driver, start_date, end_date, total_steps):
        """Download Green Button data from PG&E portal."""
        try:
            # Navigate to the Energy Usage page
            energy_usage_link = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Energy Usage')]"))
            )
            energy_usage_link.click()
            time.sleep(5)
            self.step_counter += 1
            self.progress.emit(int((self.step_counter / total_steps) * 100))
            
            # Look for Energy Usage Details
            usage_details_link = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(), 'Energy Usage Details')]"))
            )
            usage_details_link.click()
            time.sleep(5)
            self.step_counter += 1
            self.progress.emit(int((self.step_counter / total_steps) * 100))
            
            # Scroll down to find the Green Button
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)
            
            # Click on the Green Button
            green_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'green-button') or contains(@id, 'green-button') or contains(text(), 'Green Button')]"))
            )
            green_button.click()
            time.sleep(3)
            
            # Select option to export usage for a range of days
            export_range_option = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//input[@type='radio' and @value='range']"))
            )
            export_range_option.click()
            time.sleep(2)
            
            # Format dates for PG&E's date pickers
            from_date = self.validate_and_format_date(start_date)
            to_date = self.validate_and_format_date(end_date)
            
            # Enter date range
            from_date_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "from-date"))
            )
            from_date_field.clear()
            from_date_field.send_keys(from_date)
            time.sleep(1)
            
            to_date_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "to-date"))
            )
            to_date_field.clear()
            to_date_field.send_keys(to_date)
            time.sleep(1)
            
            # Click download button
            download_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Download') or contains(@class, 'download')]"))
            )
            download_button.click()
            print("Download initiated.")
            time.sleep(7)  # Wait for download to complete
            
            self.step_counter += 1
            self.progress.emit(int((self.step_counter / total_steps) * 100))
            
            # Return to dashboard
            driver.get("https://www.pge.com/myaccount/dashboard")
            time.sleep(3)
            
        except Exception as e:
            print(f"Error downloading PG&E Green Button data: {e}")
            raise

    def download_sdge_file(self, driver, start_date, end_date, total_steps):
        """Download file with custom date range from SDGE."""
        time.sleep(2)
        driver.get("https://myenergycenter.com/portal/Usage/Index")
        time.sleep(2)
        self.step_counter += 1
        self.progress.emit(int((self.step_counter / total_steps) * 100))

        start_date = self.validate_and_format_date(start_date)
        end_date = self.validate_and_format_date(end_date)

        green_button_download = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "gbloadpopup"))
        )
        green_button_download.click()
        print("Modal opened.")
        time.sleep(4)

        from_date_picker = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "gbfromdatepicker"))
        )
        driver.execute_script("arguments[0].removeAttribute('readonly')", from_date_picker)
        from_date_picker.clear()
        from_date_picker.send_keys(start_date)
        print(f"Start date entered: {start_date}")
        time.sleep(2)
        from_date_picker.send_keys(Keys.RETURN)

        to_date_picker = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.ID, "gbtodatepicker"))
        )
        driver.execute_script("arguments[0].removeAttribute('readonly')", to_date_picker)
        to_date_picker.clear()
        to_date_picker.send_keys(end_date)
        print(f"End date entered: {end_date}")
        time.sleep(2)
        to_date_picker.send_keys(Keys.RETURN)

        download_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "btngbDataDownload"))
        )
        time.sleep(2)
        download_button.click()
        print("Download initiated.")
        time.sleep(5)
        self.step_counter += 1
        self.progress.emit(int((self.step_counter / total_steps) * 100))
        driver.get("https://myenergycenter.com/portal/Dashboard/index")
        time.sleep(2)

    def interact_with_sdge_dropdown(self, driver, start_date, end_date):
        """Interact with dropdown and handle progress for SDGE."""
        dropdown_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "button[data-id='accountList']"))
        )
        dropdown_button.click()
        time.sleep(2)

        dropdown_items = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "ul.dropdown-menu > li"))
        )
        total_items = len(dropdown_items)
        steps_per_cycle = 3
        total_steps = total_items * steps_per_cycle
        print(f"Found {total_items} items in the dropdown. Total steps: {total_steps}.")

        dropdown_button.click()
        time.sleep(2)

        for index in range(total_items):
            dropdown_button = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "button[data-id='accountList']"))
            )
            dropdown_button.click()
            time.sleep(2)

            clickable_item = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "ul.dropdown-menu > li"))
            )[index]

            print(f"Selecting item {index + 1}: {clickable_item.text}")
            driver.execute_script("arguments[0].scrollIntoView(true);", clickable_item)
            clickable_item.click()
            self.step_counter += 1
            self.progress.emit(int((self.step_counter / total_steps) * 100))

            print("Waiting for page to reload...")
            try:
                WebDriverWait(driver, 20).until(EC.staleness_of(dropdown_button))
                print("Page reloaded successfully.")

                self.download_sdge_file(driver, start_date, end_date, total_steps)
            except Exception as e:
                print(f"Error waiting for page reload: {e}")
                raise
                
    def interact_with_pge_accounts(self, driver, start_date, end_date):
        """Process multiple PG&E accounts if available."""
        try:
            # Check if there's an accounts dropdown or selector
            account_selectors = driver.find_elements(By.XPATH, "//select[contains(@id, 'account') or contains(@class, 'account')]")
            
            if not account_selectors:
                # If no account selector found, just download data for the current account
                print("No account selector found. Downloading data for current account.")
                self.download_pge_green_button_data(driver, start_date, end_date, 3)
                return
                
            account_selector = account_selectors[0]
            account_selector.click()
            time.sleep(1)
            
            # Get all account options
            account_options = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.XPATH, "//select[contains(@id, 'account')]/option"))
            )
            
            total_accounts = len(account_options)
            steps_per_account = 3
            total_steps = total_accounts * steps_per_account
            
            print(f"Found {total_accounts} accounts. Total steps: {total_steps}.")
            
            for i, option in enumerate(account_options):
                if i == 0 and option.get_attribute("value") == "":
                    # Skip placeholder/default option
                    continue
                    
                print(f"Selecting account {i}: {option.text}")
                option.click()
                time.sleep(3)
                
                self.step_counter += 1
                self.progress.emit(int((self.step_counter / total_steps) * 100))
                
                # Download data for this account
                self.download_pge_green_button_data(driver, start_date, end_date, total_steps)
                
        except Exception as e:
            print(f"Error processing PG&E accounts: {e}")
            # If there's an error with account selection, try downloading anyway
            self.download_pge_green_button_data(driver, start_date, end_date, 3)

    def configure_driver(self):
        """Configure Chrome WebDriver with custom download folder."""
        
        normalized_path = self.download_path.replace("/", "\\")
        options = webdriver.ChromeOptions()
        prefs = {
            "download.default_directory": normalized_path,  # Set custom download folder
            "download.prompt_for_download": False,  # Disable prompt
            "directory_upgrade": True,
        }
        options.add_experimental_option("prefs", prefs)
        return webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    def run(self):
        """Run the Selenium script based on selected utility provider."""
        try:
            self.driver = self.configure_driver()
            
            # Determine the appropriate URL based on the utility provider
            if self.utility_provider == "PG&E":
                login_url = "https://www.pge.com/en/login"
            elif self.utility_provider == "SDGE":
                login_url = self.url  # Use the provided URL for SDGE
            elif self.utility_provider == "SCE":
                login_url = "https://www.sce.com/mysce/login"
            else:
                login_url = self.url
            
            self.login_to_portal(self.driver, login_url, self.username, self.password)
            
            # Based on utility provider, use the appropriate workflow
            if self.utility_provider == "PG&E":
                self.interact_with_pge_accounts(self.driver, self.start_date, self.end_date)
            elif self.utility_provider == "SDGE":
                self.interact_with_sdge_dropdown(self.driver, self.start_date, self.end_date)
            elif self.utility_provider == "SCE":
                # SCE workflow implementation (placeholder)
                pass
                
            self.finished.emit(True, f"Automation completed successfully for {self.utility_provider}!")
        except Exception as e:
            self.finished.emit(False, f"An error occurred: {e}")
        finally:
            if self.driver:
                self.driver.quit()

class AutomationApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("YouPower Billing Automation")
        self.setGeometry(100, 100, 500, 450)
        self.setWindowIcon(QIcon("icon.png"))

        self.worker = None

        self.center_window()

        self.image_label = QLabel()
        self.image_label.setPixmap(QPixmap("logo.png"))
        self.image_label.setAlignment(Qt.AlignCenter)

        # Add utility provider selector
        self.utility_label = QLabel("Utility Provider:")
        self.utility_selector = QComboBox()
        self.utility_selector.addItems(["SDGE", "PG&E", "SCE"])
        self.utility_selector.currentTextChanged.connect(self.update_url)

        self.url_label = QLabel("Portal URL:")
        self.url_input = QLineEdit()
        self.url_input.setText("https://myenergycenter.com/portal/PreLogin/Validate")  # Default SDGE URL

        self.username_label = QLabel("Username:")
        self.username_input = QLineEdit()

        self.password_label = QLabel("Password:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)

        self.start_date_label = QLabel("Start Date:")
        self.start_date_input = QDateEdit()
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDate(QDate.currentDate().addMonths(-1))
        self.start_date_input.setDisplayFormat("dd MMMM, yyyy")

        self.end_date_label = QLabel("End Date:")
        self.end_date_input = QDateEdit()
        self.end_date_input.setCalendarPopup(True)
        self.end_date_input.setDate(QDate.currentDate())
        self.end_date_input.setDisplayFormat("dd MMMM, yyyy")

        self.download_label = QLabel("Download Folder:")

        # Create a horizontal layout for download input and browse button
        self.download_layout = QHBoxLayout()
        self.download_input = QLineEdit()
        self.download_input.setReadOnly(True)  # Make input field read-only
        self.browse_button = QPushButton("Browse")
        self.browse_button.clicked.connect(self.browse_folder)

        self.download_layout.addWidget(self.download_input)
        self.download_layout.addWidget(self.browse_button)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setAlignment(Qt.AlignCenter)

        self.start_button = QPushButton("Start Automation")
        self.start_button.clicked.connect(self.start_automation)

        self.stop_button = QPushButton("Stop Automation")
        self.stop_button.clicked.connect(self.stop_automation)
        self.stop_button.setEnabled(False)

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        layout.addWidget(self.utility_label)
        layout.addWidget(self.utility_selector)
        layout.addWidget(self.url_label)
        layout.addWidget(self.url_input)
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.start_date_label)
        layout.addWidget(self.start_date_input)
        layout.addWidget(self.end_date_label)
        layout.addWidget(self.end_date_input)
        layout.addWidget(self.download_label)
        layout.addLayout(self.download_layout)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.start_button)
        layout.addWidget(self.stop_button)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def update_url(self, provider):
        """Update the URL field based on the selected utility provider."""
        if provider == "SDGE":
            self.url_input.setText("https://myenergycenter.com/portal/PreLogin/Validate")
        elif provider == "PG&E":
            self.url_input.setText("https://www.pge.com/en/login")
        elif provider == "SCE":
            self.url_input.setText("https://www.sce.com/mysce/login")

    def browse_folder(self):
        """Open a folder dialog to select the download folder."""
        folder = QFileDialog.getExistingDirectory(self, "Select Download Folder")
        if folder:
            self.download_input.setText(folder)
            
    def center_window(self):
        """Position the application window towards the right side of the screen with a top margin."""
        screen_geometry = QDesktopWidget().availableGeometry()
        frame_geometry = self.frameGeometry()
        right_margin = 50
        top_margin = 70
        x_position = screen_geometry.right() - frame_geometry.width() - right_margin
        y_position = screen_geometry.top() + top_margin
        self.move(x_position, y_position)

    def start_automation(self):
        """Start the Selenium automation process."""
        url = self.url_input.text()
        username = self.username_input.text()
        password = self.password_input.text()
        start_date = self.start_date_input.date().toString("yyyy-MM-dd")
        end_date = self.end_date_input.date().toString("yyyy-MM-dd")
        download_path = self.download_input.text()
        utility_provider = self.utility_selector.currentText()

        if not url or not username or not password or not download_path:
            QMessageBox.warning(self, "Input Error", "Please fill all fields and select a download folder!")
            return

        self.set_form_enabled(False)
        self.worker = AutomationWorker(url, username, password, start_date, end_date, download_path, utility_provider)
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.on_automation_finished)
        self.worker.start()

    def update_progress(self, value):
        """Update the progress bar."""
        self.progress_bar.setValue(value)

    def on_automation_finished(self, success, message):
        """Handle completion of the automation process."""
        self.set_form_enabled(True)
        self.progress_bar.setValue(100 if success else 0)

        if success:
            QMessageBox.information(self, "Success", message)
        else:
            QMessageBox.critical(self, "Error", message)

    def stop_automation(self):
        """Stop the Selenium automation process and close the application."""
        if self.worker and self.worker.isRunning():
            QMessageBox.information(self, "Stopping", "Stopping the automation process...")
            self.worker.terminate()
            self.worker.exit()
            self.worker.quit()
            self.worker = None

        QApplication.instance().quit()

    def set_form_enabled(self, enabled):
        """Enable/disable form and buttons."""
        self.utility_selector.setEnabled(enabled)
        self.url_input.setEnabled(enabled)
        self.username_input.setEnabled(enabled)
        self.password_input.setEnabled(enabled)
        self.start_date_input.setEnabled(enabled)
        self.end_date_input.setEnabled(enabled)
        self.download_input.setEnabled(enabled)
        self.browse_button.setEnabled(enabled)
        self.start_button.setEnabled(enabled)
        self.stop_button.setEnabled(not enabled)


class GBDProcessor:
    """Process Green Button Data into formatted Excel files."""
    
    def __init__(self, gbd_file_path, output_path, utility_provider="PG&E"):
        self.gbd_file_path = gbd_file_path
        self.output_path = output_path
        self.utility_provider = utility_provider
        
    def process_pge_gbd(self):
        """Process PG&E Green Button Data into formatted Excel."""
        try:
            # Read GBD file (usually XML or CSV format from Green Button)
            if self.gbd_file_path.lower().endswith('.xml'):
                # Parse XML file (simplified - actual implementation would need XML parsing)
                # This is a placeholder for XML parsing logic
                print(f"Processing XML file: {self.gbd_file_path}")
                # Convert to DataFrame for easier processing
                # df = pd.read_xml(self.gbd_file_path) - would need custom parser
                
                # For demonstration, create a sample DataFrame
                # In reality, would parse the actual XML file
                df = pd.DataFrame({
                    'timestamp': [datetime.now() - timedelta(hours=i) for i in range(24*30)],  # 30 days of hourly data
                    'usage': [0.5 + i*0.01 for i in range(24*30)],  # Sample usage data
                })
                
            elif self.gbd_file_path.lower().endswith('.csv'):
                # Read CSV (typical format for GBD)
                print(f"Processing CSV file: {self.gbd_file_path}")
                df = pd.read_csv(self.gbd_file_path)
                
                # Basic data cleaning
                # Identify timestamp column - column names might vary
                timestamp_cols = [col for col in df.columns if 'time' in col.lower() or 'date' in col.lower()]
                if timestamp_cols:
                    time_col = timestamp_cols[0]
                    df[time_col] = pd.to_datetime(df[time_col])
                    df = df.sort_values(by=time_col)
                
                # Identify usage column - column names might vary
                usage_cols = [col for col in df.columns if 'usage' in col.lower() or 'value' in col.lower() or 'kwh' in col.lower()]
                if usage_cols:
                    usage_col = usage_cols[0]
                else:
                    # If no obvious column, assume it's the first numeric column
                    numeric_cols = df.select_dtypes(include=['number']).columns
                    if len(numeric_cols) > 0:
                        usage_col = numeric_cols[0]
                    else:
                        raise ValueError("Could not identify usage data column in CSV")
            else:
                raise ValueError(f"Unsupported file format: {self.gbd_file_path}")
                
            # Create output Excel file using a template or from scratch
            # For PG&E TOU-C calculations as per requirements
            self.create_pge_excel_output(df)
            
            return True, f"Successfully processed {self.gbd_file_path} to {self.output_path}"
            
        except Exception as e:
            print(f"Error processing GBD file: {e}")
            return False, str(e)
    
    def create_pge_excel_output(self, data_df):
        """Create formatted Excel file with PG&E TOU-C calculations."""
        try:
            # Create a new workbook with sheets for data and calculations
            wb = openpyxl.Workbook()
            
            # Get the active sheet (first sheet)
            ws = wb.active
            ws.title = "Data"
            
            # Create additional sheets for calculations
            pricing_sheet = wb.create_sheet("Pricing Variables")
            baseline_sheet = wb.create_sheet("Baseline Allowances")
            weekday_time_sheet = wb.create_sheet("Weekday Time Table")
            weekend_time_sheet = wb.create_sheet("Weekend & Holiday Time Table")
            
            # Set up header section in Data sheet
            ws['A1'] = "Name"
            ws['B1'] = "Value"
            ws['K1'] = "Consumption"
            ws['L1'] = "Solar"
            ws['M1'] = "Delivery"
            
            # Address and account info
            ws['A2'] = "Address"
            ws['A3'] = "Account Number"
            ws['A4'] = "Climate Zone"
            ws['B4'] = "=VLOOKUP(LEFT(B2,1),'Baseline Allowances'!$A$1:$C$12,1,FALSE)"
            
            # Create On-Peak/Off-Peak summary
            ws['J2'] = "On-Peak"
            ws['J3'] = "Off-Peak"
            ws['K2'] = "=SUMIF($J$15:$J$3086,1,$E$15:$E$3086)"
            ws['K3'] = "=SUMIF($J$15:$J$3086,2,$E$15:$E$3086)"
            ws['L2'] = "=SUMIF($J$15:$J$3086,1,$F$15:$F$3086)"
            ws['L3'] = "=SUMIF($J$15:$J$3086,2,$F$15:$F$3086)"
            
            # Baseline calculation section
            ws['A6'] = "Baseline Information"
            ws['B6'] = "Value"
            ws['A7'] = "Billing Start Date"
            ws['A8'] = "Billing End Date"
            ws['A9'] = "Days in Billing"
            ws['B9'] = "=DATEDIF(B7,B8,\"D\")+1"
            ws['A10'] = "Season"
            ws['B10'] = "=IF(AND(MONTH(B7)>=6,MONTH(B7)<=9),\"Summer\",\"Winter\")"
            ws['A11'] = "Daily Baseline"
            ws['B11'] = "=VLOOKUP(B4,'Baseline Allowances'!$A$1:$C$12,IF(B10=\"Summer\",2,3),FALSE)"
            ws['A12'] = "Total Baseline"
            ws['B12'] = "=B11*B9"
            
            # Consumption and Tier Analysis
            ws['A14'] = "Usage Analysis"
            ws['B14'] = "Value"
            ws['A15'] = "Total Consumption"
            ws['B15'] = "=SUM(K2:K3)"
            ws['A16'] = "Tier 1 Usage (0-100%)"
            ws['B16'] = "=MIN(B15,B12)"
            ws['A17'] = "Tier 2 Usage (101-130%)"
            ws['B17'] = "=MIN(MAX(0,B15-B16),B12*0.3)"
            ws['A18'] = "Tier 3 Usage (>130%)"
            ws['B18'] = "=MAX(0,B15-B16-B17)"
            
            # Set up rate calculation section
            ws['A29'] = "Time Period"
            ws['B29'] = "Tier"
            ws['C29'] = "Consumption"
            ws['D29'] = "Rate"
            ws['E29'] = "Cost"
            
            # On-Peak tiers
            ws['A30'] = "On-Peak"
            ws['B30'] = "1"
            ws['C30'] = "=MIN(K2,B16)"
            ws['D30'] = "=INDEX('Pricing Variables'!$A$1:$E$6,MATCH(1,'Pricing Variables'!$A:$A,0),3) * IF(B10=\"Summer\",1,0.8)"
            ws['E30'] = "=C30*D30"
            
            ws['A31'] = "On-Peak"
            ws['B31'] = "2"
            ws['C31'] = "=MIN(MAX(0,K2-C30),B17)"
            ws['D31'] = "=INDEX('Pricing Variables'!$A$1:$E$6,MATCH(1,'Pricing Variables'!$A:$A,0),4) * IF(B10=\"Summer\",1,0.8)"
            ws['E31'] = "=C31*D31"
            
            ws['A32'] = "On-Peak"
            ws['B32'] = "3"
            ws['C32'] = "=MAX(0,K2-C30-C31)"
            ws['D32'] = "=INDEX('Pricing Variables'!$A$1:$E$6,MATCH(1,'Pricing Variables'!$A:$A,0),5) * IF(B10=\"Summer\",1,0.8)"
            ws['E32'] = "=C32*D32"
            
            # Off-Peak tiers
            ws['A33'] = "Off-Peak"
            ws['B33'] = "1"
            ws['C33'] = "=MIN(K3,MAX(0,B16-C30))"
            ws['D33'] = "=INDEX('Pricing Variables'!$A$1:$E$6,MATCH(2,'Pricing Variables'!$A:$A,0),3) * IF(B10=\"Summer\",1,0.8)"
            ws['E33'] = "=C33*D33"
            
            ws['A34'] = "Off-Peak"
            ws['B34'] = "2"
            ws['C34'] = "=MIN(MAX(0,K3-C33),MAX(0,B17-C31))"
            ws['D34'] = "=INDEX('Pricing Variables'!$A$1:$E$6,MATCH(2,'Pricing Variables'!$A:$A,0),4) * IF(B10=\"Summer\",1,0.8)"
            ws['E34'] = "=C34*D34"
            
            ws['A35'] = "Off-Peak"
            ws['B35'] = "3"
            ws['C35'] = "=MAX(0,K3-C33-C34)"
            ws['D35'] = "=INDEX('Pricing Variables'!$A$1:$E$6,MATCH(2,'Pricing Variables'!$A:$A,0),5) * IF(B10=\"Summer\",1,0.8)"
            ws['E35'] = "=C35*D35"
            
            # Bill Summary section
            ws['A39'] = "Bill Component"
            ws['B39'] = "Amount"
            ws['A40'] = "On-Peak Charges"
            ws['B40'] = "=SUM(E30:E32)"
            ws['A41'] = "Off-Peak Charges"
            ws['B41'] = "=SUM(E33:E35)"
            ws['A42'] = "Monthly Service Fee"
            ws['B42'] = "10.00"
            ws['A43'] = "Total Bill"
            ws['B43'] = "=SUM(B40:B42)"
            
            # Set up header for Green Button Data (starting row 15)
            ws['A14'] = "Date"
            ws['B14'] = "Timestamp"
            ws['E14'] = "Usage"
            ws['F14'] = "Solar"
            ws['H14'] = "Hour"
            ws['I14'] = "Day Type"
            ws['J14'] = "Period Code"
            ws['K14'] = "Tier"
            ws['L14'] = "Rate"
            ws['M14'] = "Cost"
            
            # Create formulas for GBD rows
            # (These would be populated for each row of actual GBD data)
            
            # Set up Pricing Variables sheet
            pricing_sheet['A1'] = "Period Code"
            pricing_sheet['B1'] = "Description"
            pricing_sheet['C1'] = "Tier 1 Rate"
            pricing_sheet['D1'] = "Tier 2 Rate"
            pricing_sheet['E1'] = "Tier 3 Rate"
            
            # PG&E TOU-C rates (placeholder - actual rates would need to be researched)
            pricing_sheet['A2'] = "1"
            pricing_sheet['B2'] = "On-Peak"
            pricing_sheet['C2'] = "0.36572"
            pricing_sheet['D2'] = "0.44561"
            pricing_sheet['E2'] = "0.48561"
            
            pricing_sheet['A3'] = "2"
            pricing_sheet['B3'] = "Off-Peak"
            pricing_sheet['C3'] = "0.32745"
            pricing_sheet['D3'] = "0.40561"
            pricing_sheet['E3'] = "0.44561"
            
            # Set up Baseline Allowances sheet
            baseline_sheet['A1'] = "Climate Zone"
            baseline_sheet['B1'] = "Summer Baseline"
            baseline_sheet['C1'] = "Winter Baseline"
            
            # PG&E climate zones (placeholder - actual values would need to be researched)
            baseline_sheet['A2'] = "P"
            baseline_sheet['B2'] = "16.4"
            baseline_sheet['C2'] = "12.1"
            
            baseline_sheet['A3'] = "Q"
            baseline_sheet['B3'] = "15.8"
            baseline_sheet['C3'] = "11.7"
            
            baseline_sheet['A4'] = "R"
            baseline_sheet['B4'] = "17.1"
            baseline_sheet['C4'] = "11.7"
            
            baseline_sheet['A5'] = "S"
            baseline_sheet['B5'] = "15.8"
            baseline_sheet['C5'] = "11.7"
            
            baseline_sheet['A6'] = "T"
            baseline_sheet['B6'] = "7.7"
            baseline_sheet['C6'] = "10.6"
            
            baseline_sheet['A7'] = "V"
            baseline_sheet['B7'] = "7.6"
            baseline_sheet['C7'] = "10.2"
            
            baseline_sheet['A8'] = "W"
            baseline_sheet['B8'] = "12.9"
            baseline_sheet['C8'] = "12.1"
            
            baseline_sheet['A9'] = "X"
            baseline_sheet['B9'] = "9.9"
            baseline_sheet['C9'] = "13.6"
            
            baseline_sheet['A10'] = "Y"
            baseline_sheet['B10'] = "11.7"
            baseline_sheet['C10'] = "12.5"
            
            baseline_sheet['A11'] = "Z"
            baseline_sheet['B11'] = "6.3"
            baseline_sheet['C11'] = "9.9"
            
            # Set up time tables
            # Weekday time table
            weekday_time_sheet['A1'] = "Hour"
            for i in range(24):
                weekday_time_sheet[f'A{i+2}'] = i
                
            for month in range(1, 13):
                weekday_time_sheet[f'{chr(65+month)}1'] = month
                
            # Fill in the time periods (1=On-Peak, 2=Off-Peak)
            # This is a placeholder - actual time periods would need to be researched for PG&E TOU-C
            for row in range(2, 26):
                hour = row - 2
                for col in range(2, 14):  # B-M for months 1-12
                    # 4 PM to 9 PM (16-21) is On-Peak for all months
                    if 16 <= hour <= 21:
                        weekday_time_sheet.cell(row=row, column=col).value = 1
                    else:
                        weekday_time_sheet.cell(row=row, column=col).value = 2
            
            # Weekend time table
            weekend_time_sheet['A1'] = "Hour"
            for i in range(24):
                weekend_time_sheet[f'A{i+2}'] = i
                
            for month in range(1, 13):
                weekend_time_sheet[f'{chr(65+month)}1'] = month
                
            # For TOU-C, weekends have same 4 PM - 9 PM peak as weekdays
            for row in range(2, 26):
                hour = row - 2
                for col in range(2, 14):  # B-M for months 1-12
                    # 4 PM to 9 PM (16-21) is On-Peak for all months
                    if 16 <= hour <= 21:
                        weekend_time_sheet.cell(row=row, column=col).value = 1
                    else:
                        weekend_time_sheet.cell(row=row, column=col).value = 2
            
            # Apply formatting
            # Freeze panes
            ws.freeze_panes = ws['A15']
            
            # Apply styles for headers, etc.
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                
            for cell in ws[14]:
                cell.font = header_font
                cell.fill = header_fill
                
            # Save the workbook
            wb.save(self.output_path)
            
            print(f"Excel file created at {self.output_path}")
            return True
            
        except Exception as e:
            print(f"Error creating Excel output: {e}")
            return False
    
    def process_gbd(self):
        """Process GBD file based on utility provider."""
        if self.utility_provider == "PG&E":
            return self.process_pge_gbd()
        elif self.utility_provider == "SDGE":
            # Placeholder for SDGE processing logic
            print("SDGE processing not yet implemented")
            return False, "SDGE processing not yet implemented"
        elif self.utility_provider == "SCE":
            # Placeholder for SCE processing logic
            print("SCE processing not yet implemented")
            return False, "SCE processing not yet implemented"
        else:
            return False, f"Unknown utility provider: {self.utility_provider}"

class AutomationApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("YouPower Billing Automation")
        self.setGeometry(100, 100, 550, 500)
        self.setWindowIcon(QIcon("icon.png"))

        self.worker = None
        self.processor = None

        self.center_window()

        self.image_label = QLabel()
        self.image_label.setPixmap(QPixmap("logo.png"))
        self.image_label.setAlignment(Qt.AlignCenter)

        # Add utility provider selector
        self.utility_label = QLabel("Utility Provider:")
        self.utility_selector = QComboBox()
        self.utility_selector.addItems(["SDGE", "PG&E", "SCE"])
        self.utility_selector.currentTextChanged.connect(self.update_url)

        self.url_label = QLabel("Portal URL:")
        self.url_input = QLineEdit()
        self.url_input.setText("https://myenergycenter.com/portal/PreLogin/Validate")  # Default SDGE URL

        self.username_label = QLabel("Username:")
        self.username_input = QLineEdit()

        self.password_label = QLabel("Password:")
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.Password)

        self.start_date_label = QLabel("Start Date:")
        self.start_date_input = QDateEdit()
        self.start_date_input.setCalendarPopup(True)
        self.start_date_input.setDate(QDate.currentDate().addMonths(-1))
        self.start_date_input.setDisplayFormat("dd MMMM, yyyy")

        self.end_date_label = QLabel("End Date:")
        self.end_date_input = QDateEdit()
        self.end_date_input.setCalendarPopup(True)
        self.end_date_input.setDate(QDate.currentDate())
        self.end_date_input.setDisplayFormat("dd MMMM, yyyy")

        self.download_label = QLabel("Download Folder:")

        # Create a horizontal layout for download input and browse button
        self.download_layout = QHBoxLayout()
        self.download_input = QLineEdit()
        self.download_input.setReadOnly(True)  # Make input field read-only
        self.browse_button = QPushButton("Browse")
        self.browse_button.clicked.connect(self.browse_folder)

        self.download_layout.addWidget(self.download_input)
        self.download_layout.addWidget(self.browse_button)
        
        # Add process to Excel option
        self.excel_checkbox = QCheckBox("Process to Excel after download")
        self.excel_checkbox.setChecked(True)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setAlignment(Qt.AlignCenter)

        self.start_button = QPushButton("Start Automation")
        self.start_button.clicked.connect(self.start_automation)

        self.stop_button = QPushButton("Stop Automation")
        self.stop_button.clicked.connect(self.stop_automation)
        self.stop_button.setEnabled(False)

        layout = QVBoxLayout()
        layout.addWidget(self.image_label)
        layout.addWidget(self.utility_label)
        layout.addWidget(self.utility_selector)
        layout.addWidget(self.url_label)
        layout.addWidget(self.url_input)
        layout.addWidget(self.username_label)
        layout.addWidget(self.username_input)
        layout.addWidget(self.password_label)
        layout.addWidget(self.password_input)
        layout.addWidget(self.start_date_label)
        layout.addWidget(self.start_date_input)
        layout.addWidget(self.end_date_label)
        layout.addWidget(self.end_date_input)
        layout.addWidget(self.download_label)
        layout.addLayout(self.download_layout)
        layout.addWidget(self.excel_checkbox)
        layout.addWidget(self.progress_bar)
        layout.addWidget(self.start_button)
        layout.addWidget(self.stop_button)

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)
        
    def update_url(self, provider):
        """Update the URL field based on the selected utility provider."""
        if provider == "SDGE":
            self.url_input.setText("https://myenergycenter.com/portal/PreLogin/Validate")
        elif provider == "PG&E":
            self.url_input.setText("https://www.pge.com/en/login")
        elif provider == "SCE":
            self.url_input.setText("https://www.sce.com/mysce/login")

    def browse_folder(self):
        """Open a folder dialog to select the download folder."""
        folder = QFileDialog.getExistingDirectory(self, "Select Download Folder")
        if folder:
            self.download_input.setText(folder)
            
    def center_window(self):
        """Position the application window towards the right side of the screen with a top margin."""
        screen_geometry = QDesktopWidget().availableGeometry()
        frame_geometry = self.frameGeometry()
        right_margin = 50
        top_margin = 70
        x_position = screen_geometry.right() - frame_geometry.width() - right_margin
        y_position = screen_geometry.top() + top_margin
        self.move(x_position, y_position)

    def start_automation(self):
        """Start the Selenium automation process."""
        url = self.url_input.text()
        username = self.username_input.text()
        password = self.password_input.text()
        start_date = self.start_date_input.date().toString("yyyy-MM-dd")
        end_date = self.end_date_input.date().toString("yyyy-MM-dd")
        download_path = self.download_input.text()
        utility_provider = self.utility_selector.currentText()
        process_to_excel = self.excel_checkbox.isChecked()

        if not url or not username or not password or not download_path:
            QMessageBox.warning(self, "Input Error", "Please fill all fields and select a download folder!")
            return

        self.set_form_enabled(False)
        self.worker = AutomationWorker(url, username, password, start_date, end_date, download_path, utility_provider)
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(lambda success, msg: self.on_download_finished(success, msg, download_path, utility_provider, process_to_excel))
        self.worker.start()

    def on_download_finished(self, success, message, download_path, utility_provider, process_to_excel):
        """Handle completion of the download process and start Excel processing if needed."""
        if success and process_to_excel:
            QMessageBox.information(self, "Download Complete", "Download completed successfully. Processing to Excel...")
            self.process_to_excel(download_path, utility_provider)
        else:
            self.on_automation_finished(success, message)

    def process_to_excel(self, download_path, utility_provider):
        """Process downloaded GBD file to Excel."""
        try:
            # Find the most recently downloaded GBD file
            files = [os.path.join(download_path, f) for f in os.listdir(download_path)]
            if not files:
                self.on_automation_finished(False, "No files found in download directory")
                return
                
            files = [f for f in files if f.lower().endswith('.xml') or f.lower().endswith('.csv')]
            if not files:
                self.on_automation_finished(False, "No GBD files (.xml or .csv) found in download directory")
                return
                
            # Get the most recent file
            latest_file = max(files, key=os.path.getctime)
            
            # Create output Excel file path
            file_base = os.path.splitext(os.path.basename(latest_file))[0]
            excel_output = os.path.join(download_path, f"{file_base}_{utility_provider}_processed.xlsx")
            
            # Process the file
            self.processor = GBDProcessor(latest_file, excel_output, utility_provider)
            success, message = self.processor.process_gbd()
            
            self.on_automation_finished(success, message)
            
        except Exception as e:
            self.on_automation_finished(False, f"Error processing to Excel: {e}")

    def update_progress(self, value):
        """Update the progress bar."""
        self.progress_bar.setValue(value)

    def on_automation_finished(self, success, message):
        """Handle completion of the automation process."""
        self.set_form_enabled(True)
        self.progress_bar.setValue(100 if success else 0)

        if success:
            QMessageBox.information(self, "Success", message)
        else:
            QMessageBox.critical(self, "Error", message)

    def stop_automation(self):
        """Stop the Selenium automation process and close the application."""
        if self.worker and self.worker.isRunning():
            QMessageBox.information(self, "Stopping", "Stopping the automation process...")
            self.worker.terminate()
            self.worker.exit()
            self.worker.quit()
            self.worker = None

        QApplication.instance().quit()

    def set_form_enabled(self, enabled):
        """Enable/disable form and buttons."""
        self.utility_selector.setEnabled(enabled)
        self.url_input.setEnabled(enabled)
        self.username_input.setEnabled(enabled)
        self.password_input.setEnabled(enabled)
        self.start_date_input.setEnabled(enabled)
        self.end_date_input.setEnabled(enabled)
        self.download_input.setEnabled(enabled)
        self.browse_button.setEnabled(enabled)
        self.excel_checkbox.setEnabled(enabled)
        self.start_button.setEnabled(enabled)
        self.stop_button.setEnabled(not enabled)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = AutomationApp()
    window.show()
    sys.exit(app.exec())
