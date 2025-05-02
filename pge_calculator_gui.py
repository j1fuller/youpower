# pge_calculator_gui.py
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

class PGECalculator:
    # [Insert your existing PGECalculator class code here]
    # This is the same code you already have
    
    def __init__(self):
        # Initialize rate data and configuration
        self.initialize_rates()
        self.initialize_time_tables()
        self.initialize_baseline_allowances()
        
    def initialize_rates(self):
        # Pricing variables equivalent to Excel sheet
        self.rates = {
            "on_peak": {
                "tier1": 0.45,
                "tier2": 0.55,
                "tier3": 0.65
            },
            "off_peak": {
                "tier1": 0.32,
                "tier2": 0.42,
                "tier3": 0.52
            }
        }
        
        # Define holidays
        self.holidays = [
            "2025-01-01", "2025-02-17", "2025-05-26", "2025-07-04", 
            "2025-09-01", "2025-11-11", "2025-11-27", "2025-12-25"
        ]
        self.holidays = [datetime.strptime(d, "%Y-%m-%d") for d in self.holidays]
        
        # Monthly service fee
        self.monthly_service_fee = 10.00
        
    def initialize_time_tables(self):
        # Simple rules for PG&E TOU-C
        pass
        
    def initialize_baseline_allowances(self):
        # Baseline allowances by climate zone
        self.baseline_allowances = {
            "P": {"summer": 15.8, "winter": 12.9},
            "Q": {"summer": 7.7, "winter": 10.9},
            "R": {"summer": 18.9, "winter": 11.7},
            "S": {"summer": 15.8, "winter": 12.0},
            "T": {"summer": 7.6, "winter": 9.7},
            "V": {"summer": 8.3, "winter": 9.8},
            "W": {"summer": 20.9, "winter": 13.7},
            "X": {"summer": 11.2, "winter": 11.3},
            "Y": {"summer": 11.9, "winter": 12.6},
            "Z": {"summer": 7.9, "winter": 10.4}
        }
    
    def determine_time_period(self, timestamp):
        """Determine the time period (on_peak or off_peak) for a given timestamp"""
        hour = timestamp.hour
        
        # Check if it's a holiday
        for holiday in self.holidays:
            if timestamp.date() == holiday.date():
                return "off_peak"
        
        # 4-9 PM (hours 16-20) are on-peak every day
        if 16 <= hour <= 20:
            return "on_peak"
        else:
            return "off_peak"
    
    def determine_season(self, date):
        """Determine if a date is in summer or winter season"""
        month = date.month
        if 6 <= month <= 9:
            return "summer"
        else:
            return "winter"
    
    def calculate_baseline_allowance(self, climate_zone, start_date, end_date):
        """Calculate baseline allowance for a billing period"""
        # [rest of your method code]
        days = (end_date - start_date).days + 1
        
        # If billing period spans both seasons, calculate weighted average
        if self.determine_season(start_date) != self.determine_season(end_date):
            # Find the season transition date (either June 1 or October 1)
            if start_date.month < 6 and end_date.month >= 6:
                transition_date = datetime(start_date.year, 6, 1)
            else:
                transition_date = datetime(start_date.year, 10, 1)
            
            days_in_first_season = (transition_date - start_date).days
            days_in_second_season = (end_date - transition_date).days + 1
            
            first_season = self.determine_season(start_date)
            second_season = self.determine_season(end_date)
            
            allowance = (
                self.baseline_allowances[climate_zone][first_season] * days_in_first_season +
                self.baseline_allowances[climate_zone][second_season] * days_in_second_season
            )
            return allowance
        else:
            season = self.determine_season(start_date)
            return self.baseline_allowances[climate_zone][season] * days
    
    def determine_tier(self, usage, baseline_allowance):
        """Determine which tier(s) the usage falls into"""
        # [rest of your method code]
        tier1_usage = min(usage, baseline_allowance)
        tier2_usage = min(max(0, usage - baseline_allowance), baseline_allowance * 0.3)
        tier3_usage = max(0, usage - baseline_allowance * 1.3)
        
        return {
            "tier1": tier1_usage,
            "tier2": tier2_usage,
            "tier3": tier3_usage
        }
    
    def calculate_cost(self, time_period, tier, usage, season):
        """Calculate cost for a specific time period, tier, and usage amount"""
        # [rest of your method code]
        rate = self.rates[time_period][tier]
        
        # Apply seasonal adjustment for winter (80% of summer rates)
        if season == "winter":
            rate *= 0.8
            
        return rate * usage
    
    def process_gbd_data(self, gbd_file, climate_zone, start_date, end_date):
        """Process Green Button Data and calculate bill"""
        # [rest of your method code]
        # Load GBD data from CSV file
        df = pd.read_csv(gbd_file)
        
        # Convert timestamp strings to datetime objects
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        
        # Filter for the billing period
        df = df[(df['timestamp'] >= start_date) & (df['timestamp'] <= end_date)]
        
        # Add time period column
        df['time_period'] = df['timestamp'].apply(self.determine_time_period)
        
        # Summarize consumption by time period
        consumption_summary = df.groupby('time_period')['usage'].sum()
        
        # Calculate baseline allowance
        baseline_allowance = self.calculate_baseline_allowance(climate_zone, start_date, end_date)
        
        # Calculate total consumption
        total_consumption = consumption_summary.sum()
        
        # Determine tiers
        tier_usage = self.determine_tier(total_consumption, baseline_allowance)
        
        # Calculate costs for each time period and tier
        costs = []
        
        # Season for rate calculation
        season = self.determine_season(start_date) if self.determine_season(start_date) == self.determine_season(end_date) else "mixed"
        
        # [rest of your cost calculation code]
        # Calculate On-Peak costs
        on_peak_usage = consumption_summary.get('on_peak', 0)
        remaining_on_peak = on_peak_usage
        
        # Tier 1 On-Peak
        on_peak_tier1 = min(remaining_on_peak, tier_usage['tier1'])
        if on_peak_tier1 > 0:
            costs.append(self.calculate_cost('on_peak', 'tier1', on_peak_tier1, season))
            remaining_on_peak -= on_peak_tier1
        
        # Tier 2 On-Peak
        on_peak_tier2 = min(remaining_on_peak, tier_usage['tier2'])
        if on_peak_tier2 > 0:
            costs.append(self.calculate_cost('on_peak', 'tier2', on_peak_tier2, season))
            remaining_on_peak -= on_peak_tier2
        
        # Tier 3 On-Peak
        if remaining_on_peak > 0:
            costs.append(self.calculate_cost('on_peak', 'tier3', remaining_on_peak, season))
        
        # Calculate Off-Peak costs
        off_peak_usage = consumption_summary.get('off_peak', 0)
        remaining_off_peak = off_peak_usage
        
        # Tier 1 Off-Peak
        off_peak_tier1 = min(remaining_off_peak, max(0, tier_usage['tier1'] - on_peak_tier1))
        if off_peak_tier1 > 0:
            costs.append(self.calculate_cost('off_peak', 'tier1', off_peak_tier1, season))
            remaining_off_peak -= off_peak_tier1
        
        # Tier 2 Off-Peak
        off_peak_tier2 = min(remaining_off_peak, max(0, tier_usage['tier2'] - on_peak_tier2))
        if off_peak_tier2 > 0:
            costs.append(self.calculate_cost('off_peak', 'tier2', off_peak_tier2, season))
            remaining_off_peak -= off_peak_tier2
        
        # Tier 3 Off-Peak
        if remaining_off_peak > 0:
            costs.append(self.calculate_cost('off_peak', 'tier3', remaining_off_peak, season))
        
        # Add monthly service fee
        total_bill = sum(costs) + self.monthly_service_fee
        
        # Prepare result data - simplified for the GUI version
        return {
            'consumption': {
                'on_peak': consumption_summary.get('on_peak', 0),
                'off_peak': consumption_summary.get('off_peak', 0),
                'total': total_consumption
            },
            'baseline_allowance': baseline_allowance,
            'tier_usage': tier_usage,
            'monthly_service_fee': self.monthly_service_fee,
            'total_bill': total_bill
        }

def export_to_excel(calculation_results, output_file="pge_bill_calculation.xlsx"):
    """Export PG&E calculator results to Excel"""
    # [Include the Excel export function code here]
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "PG&E Bill Calculation"
    
    # Add header
    ws['A1'] = "PG&E TOU-C Bill Calculation"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Account Information
    ws['A3'] = "Account Information"
    ws['A3'].font = Font(bold=True)
    
    ws['A4'] = "Climate Zone"
    ws['B4'] = calculation_results.get('climate_zone', 'X')
    ws['A5'] = "Billing Period"
    ws['B5'] = f"{calculation_results.get('start_date', '')} to {calculation_results.get('end_date', '')}"
    ws['A6'] = "Days in Billing"
    ws['B6'] = calculation_results.get('days_in_billing', 30)
    
    # Consumption Summary
    ws['A8'] = "Consumption Summary"
    ws['A8'].font = Font(bold=True)
    
    ws['A9'] = "Time Period"
    ws['B9'] = "Usage (kWh)"
    
    ws['A10'] = "On-Peak"
    ws['B10'] = calculation_results['consumption']['on_peak']
    ws['A11'] = "Off-Peak"
    ws['B11'] = calculation_results['consumption']['off_peak']
    ws['A12'] = "Total"
    ws['B12'] = calculation_results['consumption']['total']
    
    # Baseline Information
    ws['A14'] = "Baseline Information"
    ws['A14'].font = Font(bold=True)
    
    ws['A15'] = "Baseline Allowance (kWh)"
    ws['B15'] = calculation_results['baseline_allowance']
    
    # Tier Usage
    ws['A17'] = "Tier Usage"
    ws['A17'].font = Font(bold=True)
    
    ws['A18'] = "Tier"
    ws['B18'] = "Usage (kWh)"
    
    ws['A19'] = "Tier 1 (0-100%)"
    ws['B19'] = calculation_results['tier_usage']['tier1']
    ws['A20'] = "Tier 2 (101-130%)"
    ws['B20'] = calculation_results['tier_usage']['tier2']
    ws['A21'] = "Tier 3 (>130%)"
    ws['B21'] = calculation_results['tier_usage']['tier3']
    
    # Bill Summary
    ws['A23'] = "Bill Summary"
    ws['A23'].font = Font(bold=True)
    
    ws['A24'] = "Monthly Service Fee"
    ws['B24'] = calculation_results['monthly_service_fee']
    
    ws['A25'] = "Total Bill"
    ws['B25'] = calculation_results['total_bill']
    ws['B25'].font = Font(bold=True)
    
    # Save the workbook
    wb.save(output_file)
    return output_file

def create_sample_gbd_data(output_file="sample_gbd.csv"):
    """Create a sample Green Button Data file"""
    # Create a sample dataset
    start_date = datetime(2025, 2, 1)
    end_date = datetime(2025, 2, 28)
    date_range = pd.date_range(start=start_date, end=end_date, freq='H')
    
    # Generate some realistic usage patterns with higher usage during peak hours
    usage = []
    for dt in date_range:
        hour = dt.hour
        # Higher usage during evening peak hours
        if 16 <= hour <= 20:
            base_usage = np.random.uniform(1.2, 2.5)  # Higher usage during peak
        else:
            base_usage = np.random.uniform(0.3, 1.0)  # Lower usage off-peak
        usage.append(base_usage)
    
    # Create dataframe
    data = {
        'timestamp': date_range,
        'usage': usage
    }
    df = pd.DataFrame(data)
    
    # Save to CSV
    df.to_csv(output_file, index=False)
    return output_file

# Create GUI application
class PGECalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("PG&E TOU-C Calculator")
        self.root.geometry("600x600")
        self.root.resizable(True, True)
        
        # Create frame for inputs
        input_frame = ttk.LabelFrame(root, text="Inputs")
        input_frame.pack(padx=10, pady=10, fill="x")
        
        # File input
        ttk.Label(input_frame, text="Green Button Data File:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
        self.file_frame = ttk.Frame(input_frame)
        self.file_frame.grid(row=0, column=1, sticky="ew", padx=5, pady=5)
        
        self.file_path = tk.StringVar()
        ttk.Entry(self.file_frame, textvariable=self.file_path, width=40).pack(side="left", padx=5)
        ttk.Button(self.file_frame, text="Browse", command=self.browse_file).pack(side="left", padx=5)
        ttk.Button(self.file_frame, text="Create Sample", command=self.create_sample).pack(side="left", padx=5)
        
        # Climate zone
        ttk.Label(input_frame, text="Climate Zone:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
        self.climate_zone = tk.StringVar(value="X")
        zones = ['P', 'Q', 'R', 'S', 'T', 'V', 'W', 'X', 'Y', 'Z']
        ttk.Combobox(input_frame, textvariable=self.climate_zone, values=zones, width=5).grid(row=1, column=1, sticky="w", padx=5, pady=5)
        
        # Date range
        ttk.Label(input_frame, text="Billing Start Date (YYYY-MM-DD):").grid(row=2, column=0, sticky="w", padx=5, pady=5)
        self.start_date = tk.StringVar(value="2025-02-01")
        ttk.Entry(input_frame, textvariable=self.start_date, width=12).grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        ttk.Label(input_frame, text="Billing End Date (YYYY-MM-DD):").grid(row=3, column=0, sticky="w", padx=5, pady=5)
        self.end_date = tk.StringVar(value="2025-02-28")
        ttk.Entry(input_frame, textvariable=self.end_date, width=12).grid(row=3, column=1, sticky="w", padx=5, pady=5)
        
        # Buttons
        button_frame = ttk.Frame(root)
        button_frame.pack(padx=10, pady=10, fill="x")
        
        ttk.Button(button_frame, text="Calculate", command=self.calculate).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Export to Excel", command=self.export).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Exit", command=root.destroy).pack(side="right", padx=5)
        
        # Results area
        result_frame = ttk.LabelFrame(root, text="Results")
        result_frame.pack(padx=10, pady=10, fill="both", expand=True)
        
        self.result_text = tk.Text(result_frame, height=20, width=70)
        self.result_text.pack(padx=5, pady=5, fill="both", expand=True)
        
        # Initialize calculator
        self.calculator = PGECalculator()
        self.result = None
        
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Green Button Data File",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")]
        )
        if file_path:
            self.file_path.set(file_path)
    
    def create_sample(self):
        try:
            file_path = create_sample_gbd_data()
            self.file_path.set(file_path)
            messagebox.showinfo("Success", f"Sample GBD data created at {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create sample data: {str(e)}")
    
    def calculate(self):
        try:
            # Validate inputs
            if not self.file_path.get():
                messagebox.showerror("Error", "Please select a GBD file")
                return
                
            # Parse dates
            try:
                start_date = datetime.strptime(self.start_date.get(), "%Y-%m-%d")
                end_date = datetime.strptime(self.end_date.get(), "%Y-%m-%d")
            except ValueError:
                messagebox.showerror("Error", "Invalid date format. Use YYYY-MM-DD")
                return
            
            # Run calculation
            self.result = self.calculator.process_gbd_data(
                self.file_path.get(),
                self.climate_zone.get(),
                start_date,
                end_date
            )
            
            # Add metadata
            self.result['climate_zone'] = self.climate_zone.get()
            self.result['start_date'] = self.start_date.get()
            self.result['end_date'] = self.end_date.get()
            self.result['days_in_billing'] = (end_date - start_date).days + 1
            
            # Display results
            self.display_results()
            
        except Exception as e:
            messagebox.showerror("Error", f"Calculation failed: {str(e)}")
    
    def display_results(self):
        # Clear previous results
        self.result_text.delete(1.0, tk.END)
        
        # Format results
        self.result_text.insert(tk.END, "PG&E TOU-C Bill Calculation\n")
        self.result_text.insert(tk.END, "=========================\n\n")
        
        self.result_text.insert(tk.END, f"Climate Zone: {self.result['climate_zone']}\n")
        self.result_text.insert(tk.END, f"Billing Period: {self.result['start_date']} to {self.result['end_date']} ({self.result['days_in_billing']} days)\n")
        self.result_text.insert(tk.END, f"Baseline Allowance: {self.result['baseline_allowance']:.2f} kWh\n\n")
        
        self.result_text.insert(tk.END, "Consumption Summary:\n")
        self.result_text.insert(tk.END, f"  On-Peak: {self.result['consumption']['on_peak']:.2f} kWh\n")
        self.result_text.insert(tk.END, f"  Off-Peak: {self.result['consumption']['off_peak']:.2f} kWh\n")
        self.result_text.insert(tk.END, f"  Total: {self.result['consumption']['total']:.2f} kWh\n\n")
        
        self.result_text.insert(tk.END, "Tier Usage:\n")
        self.result_text.insert(tk.END, f"  Tier 1 (0-100%): {self.result['tier_usage']['tier1']:.2f} kWh\n")
        self.result_text.insert(tk.END, f"  Tier 2 (101-130%): {self.result['tier_usage']['tier2']:.2f} kWh\n")
        self.result_text.insert(tk.END, f"  Tier 3 (>130%): {self.result['tier_usage']['tier3']:.2f} kWh\n\n")
        
        self.result_text.insert(tk.END, "Bill Summary:\n")
        self.result_text.insert(tk.END, f"  Monthly Service Fee: ${self.result['monthly_service_fee']:.2f}\n")
        self.result_text.insert(tk.END, f"  Total Bill: ${self.result['total_bill']:.2f}\n")
    
    def export(self):
        if not self.result:
            messagebox.showerror("Error", "Please calculate first")
            return
            
        try:
            # Ask for save location
            file_path = filedialog.asksaveasfilename(
                title="Save Excel File",
                defaultextension=".xlsx",
                filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
            )
            
            if not file_path:
                return
                
            # Export to Excel
            export_to_excel(self.result, file_path)
            messagebox.showinfo("Success", f"Exported to {file_path}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Export failed: {str(e)}")

# Main entry point
if __name__ == "__main__":
    root = tk.Tk()
    app = PGECalculatorApp(root)
    root.mainloop()
